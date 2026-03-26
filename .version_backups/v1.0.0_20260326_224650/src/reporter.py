"""
Excel 报告生成模块
直播结束后自动生成详细的 Excel 分析报告
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reports')


def _header_style(ws, row, col, value, bg_color='1F4E79', font_color='FFFFFF', bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=11)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _data_style(ws, row, col, value, bg_color=None, bold=False, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    if bg_color:
        cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format
    return cell


def generate_excel_report(summary: dict, username: str) -> str:
    """
    生成 Excel 报告
    :param summary: get_session_summary() 返回的数据
    :param username: 主播用户名
    :return: 报告文件路径
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)

    session = summary.get('session', {})
    comments = summary.get('comments', [])
    gift_rank = summary.get('gift_rank', [])
    snapshots = summary.get('snapshots', [])

    # 文件名
    start_time = session.get('start_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    date_str = start_time[:10].replace('-', '')
    time_str = start_time[11:16].replace(':', '')
    filename = f"直播报告_{username}_{date_str}_{time_str}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()

    # ==================== Sheet 1: 直播概览 ====================
    ws1 = wb.active
    ws1.title = '直播概览'
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 25

    # 标题
    ws1.merge_cells('A1:D1')
    title_cell = ws1['A1']
    title_cell.value = f'TikTok 直播数据报告 — @{username}'
    title_cell.font = Font(bold=True, size=16, color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 40

    # 基础信息
    ws1.merge_cells('A2:D2')
    ws1['A2'].value = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A2'].font = Font(size=9, color='888888')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 20

    # 核心指标
    metrics = [
        ('直播账号', f'@{username}', '直播状态', session.get('status', '-')),
        ('开始时间', session.get('start_time', '-'), '结束时间', session.get('end_time', '直播中')),
        ('峰值在线人数', session.get('peak_viewers', 0), '总观看人次', session.get('total_viewers', 0)),
        ('总评论数', session.get('total_comments', 0), '总点赞数', session.get('total_likes', 0)),
        ('新增关注', session.get('new_followers', 0), '礼物总价值($)', f"${session.get('total_gift_value', 0):.2f}"),
        ('收到礼物次数', session.get('total_gifts', 0), '礼物打赏人数', len(set(g['username'] for g in gift_rank))),
    ]

    for i, (k1, v1, k2, v2) in enumerate(metrics):
        row = i + 4
        ws1.row_dimensions[row].height = 28
        _header_style(ws1, row, 1, k1, bg_color='2E75B6')
        _data_style(ws1, row, 2, v1, bold=True)
        _header_style(ws1, row, 3, k2, bg_color='2E75B6')
        _data_style(ws1, row, 4, v2, bold=True)

    # ==================== Sheet 2: 评论记录 ====================
    ws2 = wb.create_sheet('评论记录')
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 60
    ws2.column_dimensions['E'].width = 12

    ws2.row_dimensions[1].height = 28
    _header_style(ws2, 1, 1, '#')
    _header_style(ws2, 1, 2, '时间')
    _header_style(ws2, 1, 3, '用户名')
    _header_style(ws2, 1, 4, '评论内容')
    _header_style(ws2, 1, 5, '类型')

    # 从数据库重新获取全量评论
    from src.database import get_conn
    session_id = session.get('id')
    all_comments = []
    if session_id:
        conn = get_conn()
        c = conn.cursor()
        c.execute('SELECT * FROM comments WHERE session_id=? ORDER BY timestamp', (session_id,))
        all_comments = [dict(r) for r in c.fetchall()]
        conn.close()

    for i, comment in enumerate(all_comments):
        row = i + 2
        bg = 'FFF2CC' if comment.get('is_anchor') else None
        _data_style(ws2, row, 1, i + 1, bg)
        _data_style(ws2, row, 2, comment.get('timestamp', '')[-8:], bg)
        _data_style(ws2, row, 3, comment.get('username', ''), bg)
        cell = ws2.cell(row=row, column=4, value=comment.get('content', ''))
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if bg:
            cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        _data_style(ws2, row, 5, '主播' if comment.get('is_anchor') else '观众', bg)

    # ==================== Sheet 3: 礼物排行 ====================
    ws3 = wb.create_sheet('礼物排行')
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 15

    ws3.row_dimensions[1].height = 28
    _header_style(ws3, 1, 1, '排名')
    _header_style(ws3, 1, 2, '用户名')
    _header_style(ws3, 1, 3, '礼物总价值($)')
    _header_style(ws3, 1, 4, '送礼次数')

    gold_colors = ['FFD700', 'C0C0C0', 'CD7F32']
    for i, gift in enumerate(gift_rank):
        row = i + 2
        bg = gold_colors[i] if i < 3 else ('F2F2F2' if i % 2 == 0 else None)
        _data_style(ws3, row, 1, i + 1, bg, bold=(i < 3))
        _data_style(ws3, row, 2, gift.get('username', ''), bg, bold=(i < 3))
        _data_style(ws3, row, 3, f"${gift.get('total_value', 0):.2f}", bg)
        _data_style(ws3, row, 4, gift.get('cnt', 0), bg)

    # ==================== Sheet 4: 在线人数趋势 ====================
    ws4 = wb.create_sheet('数据趋势')
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15

    ws4.row_dimensions[1].height = 28
    _header_style(ws4, 1, 1, '时间')
    _header_style(ws4, 1, 2, '在线人数')
    _header_style(ws4, 1, 3, '点赞数')
    _header_style(ws4, 1, 4, '评论数')

    for i, snap in enumerate(snapshots):
        row = i + 2
        _data_style(ws4, row, 1, snap.get('timestamp', '')[-8:])
        _data_style(ws4, row, 2, snap.get('viewer_count', 0))
        _data_style(ws4, row, 3, snap.get('like_count', 0))
        _data_style(ws4, row, 4, snap.get('comment_count', 0))

    # 添加折线图
    if len(snapshots) > 1:
        chart = LineChart()
        chart.title = '直播间数据趋势'
        chart.style = 10
        chart.y_axis.title = '数量'
        chart.x_axis.title = '时间'
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=len(snapshots) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        ws4.add_chart(chart, 'F2')

    # ==================== Sheet 5: 主播话术记录 ====================
    ws5 = wb.create_sheet('主播话术')
    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 80

    ws5.merge_cells('A1:C1')
    title5 = ws5['A1']
    title5.value = f'主播话术记录 — @{username}（语音转文字）'
    title5.font = Font(bold=True, size=13, color='7B4F00')
    title5.fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
    title5.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 32

    ws5.row_dimensions[2].height = 28
    _header_style(ws5, 2, 1, '#', bg_color='B8860B', font_color='FFFFFF')
    _header_style(ws5, 2, 2, '时间', bg_color='B8860B', font_color='FFFFFF')
    _header_style(ws5, 2, 3, '话术内容', bg_color='B8860B', font_color='FFFFFF')

    # 从数据库查主播话术（is_anchor=1 的记录）
    speech_records = []
    if session_id:
        conn = get_conn()
        c = conn.cursor()
        c.execute(
            'SELECT * FROM comments WHERE session_id=? AND is_anchor=1 ORDER BY timestamp',
            (session_id,)
        )
        speech_records = [dict(r) for r in c.fetchall()]
        conn.close()

    if speech_records:
        for i, rec in enumerate(speech_records):
            row = i + 3
            ws5.row_dimensions[row].height = 22
            _data_style(ws5, row, 1, i + 1, 'FFFDE7')
            _data_style(ws5, row, 2, rec.get('timestamp', '')[-8:], 'FFFDE7')
            cell = ws5.cell(row=row, column=3, value=rec.get('content', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.fill = PatternFill(fill_type='solid', fgColor='FFFDE7')
            thin = Side(style='thin', color='CCCCCC')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws5.cell(row=len(speech_records) + 4, column=1,
                 value=f'共 {len(speech_records)} 条话术记录').font = Font(size=9, color='888888')
    else:
        ws5.cell(row=3, column=1, value='（本场直播未采集到语音转文字数据）').font = Font(size=10, color='AAAAAA')

    wb.save(filepath)
    return filepath
